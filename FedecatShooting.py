#!/usr/bin/env python3
\"\"\"Fedecat Shooting - Prototipo ejecutable en un solo fichero

Características incluidas:
- Base de datos local SQLite (sin dependencias externas).
- Importación/Exportación sencilla a Excel (.xlsx) si está disponible openpyxl.
- Plantillas Excel de ejemplo incluidas.
- Modo CLI para operar (crear competición, importar tiradores/inscripciones/resultados, exportar resultados).
- Intento de GUI con PyQt6 si está instalado; si no, el script cae a modo CLI con instrucciones.

Uso (CLI):
  python fede_shooting_prototype.py --create-templates
  python fede_shooting_prototype.py --init-db
  python fede_shooting_prototype.py --import-shooters Tiradores.xlsx
  python fede_shooting_prototype.py --import-inscriptions Inscripciones.xlsx "Mi Competición"
  python fede_shooting_prototype.py --import-results Resultados.xlsx "Mi Competición"
  python fede_shooting_prototype.py --export-results "Mi Competición" resultados_salida.xlsx
  python fede_shooting_prototype.py --run-gui    (si PyQt6 está instalado)

Dependencias opcionales para funciones avanzadas:
  pip install openpyxl requests PyQt6
\"\"\"

import os, sys, sqlite3, json, argparse, datetime, traceback

DB_PATH = os.path.join(os.path.dirname(__file__), 'fede_shooting.db')

# ------------------ Utilidades DB ------------------
def get_conn():
    conn = sqlite3.connect(DB_PATH, detect_types=sqlite3.PARSE_DECLTYPES|sqlite3.PARSE_COLNAMES)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_conn()
    cur = conn.cursor()
    # Crear tablas básicas
    cur.executescript(\"\"\"
    CREATE TABLE IF NOT EXISTS shooters (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        club TEXT,
        categoria TEXT,
        licencia TEXT,
        pais TEXT,
        dni TEXT UNIQUE
    );
    CREATE TABLE IF NOT EXISTS competitions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL UNIQUE,
        fecha_inicio TEXT,
        fecha_fin TEXT,
        lugar TEXT,
        tipo TEXT
    );
    CREATE TABLE IF NOT EXISTS squads (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        competition_id INTEGER NOT NULL,
        nombre TEXT,
        puesto INTEGER,
        FOREIGN KEY(competition_id) REFERENCES competitions(id)
    );
    CREATE TABLE IF NOT EXISTS entries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        competition_id INTEGER NOT NULL,
        shooter_id INTEGER NOT NULL,
        squad_id INTEGER,
        dorsal TEXT,
        FOREIGN KEY(competition_id) REFERENCES competitions(id),
        FOREIGN KEY(shooter_id) REFERENCES shooters(id),
        FOREIGN KEY(squad_id) REFERENCES squads(id)
    );
    CREATE TABLE IF NOT EXISTS round_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        entry_id INTEGER NOT NULL,
        round_number INTEGER NOT NULL,
        hits INTEGER DEFAULT 0,
        misses INTEGER DEFAULT 0,
        score INTEGER DEFAULT 0,
        detail TEXT,
        timestamp TEXT,
        FOREIGN KEY(entry_id) REFERENCES entries(id)
    );
    CREATE INDEX IF NOT EXISTS ix_shooters_dni ON shooters(dni);
    CREATE INDEX IF NOT EXISTS ix_entries_comp_dorsal ON entries(competition_id, dorsal);
    \"\"\")
    conn.commit()
    conn.close()
    print(\"Base de datos inicializada en:\", DB_PATH)

# ------------------ Excel I/O (openpyxl si disponible) ------------------
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    EXCEL_AVAILABLE = True
except Exception:
    EXCEL_AVAILABLE = False

def create_templates(output_dir='.'):
    # Tiradores.xlsx
    tiradores_path = os.path.join(output_dir, 'Tiradores.xlsx')
    inscripciones_path = os.path.join(output_dir, 'Inscripciones.xlsx')
    resultados_path = os.path.join(output_dir, 'Resultados.xlsx')
    if EXCEL_AVAILABLE:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tiradores"
        ws.append(["id","nombre","club","categoria","licencia","pais","dni"])
        ws.append(["","Juan Pérez","Club A","Senior","L-123","ESP","12345678A"])
        wb.save(tiradores_path)

        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Inscripciones"
        ws2.append(["competition","squad","dorsal","shooter_dni","shooter_id","nota"])
        ws2.append(["Mi Competición","Escuadra 1","1","12345678A","",""])
        wb2.save(inscripciones_path)

        wb3 = Workbook()
        ws3 = wb3.active
        ws3.title = "Resultados"
        ws3.append(["competition","squad","dorsal","round","hits","misses","score","detalle"])
        ws3.append(["Mi Competición","Escuadra 1","1","1","24","1","24","1,1,1,0,1"])
        wb3.save(resultados_path)
        print(\"Plantillas creadas:\", tiradores_path, inscripciones_path, resultados_path)
    else:
        # fallback a CSV si openpyxl no disponible
        import csv
        with open(tiradores_path.replace('.xlsx','.csv'),'w',newline='',encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(["id","nombre","club","categoria","licencia","pais","dni"])
            w.writerow(["","Juan Pérez","Club A","Senior","L-123","ESP","12345678A"])
        with open(inscripciones_path.replace('.xlsx','.csv'),'w',newline='',encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(["competition","squad","dorsal","shooter_dni","shooter_id","nota"])
            w.writerow(["Mi Competición","Escuadra 1","1","12345678A","",""])
        with open(resultados_path.replace('.xlsx','.csv'),'w',newline='',encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(["competition","squad","dorsal","round","hits","misses","score","detalle"])
            w.writerow(["Mi Competición","Escuadra 1","1","1","24","1","24","1,1,1,0,1"])
        print(\"Openpyxl no está disponible. Generadas plantillas CSV en su lugar en el directorio.\")


def import_shooters_from_excel(path):
    conn = get_conn()
    cur = conn.cursor()
    created = 0
    if EXCEL_AVAILABLE and path.lower().endswith('.xlsx'):
        wb = load_workbook(path, read_only=True)
        if "Tiradores" not in wb.sheetnames:
            print(\"Hoja 'Tiradores' no encontrada en el fichero.\")
            return 0
        ws = wb["Tiradores"]
        rows = list(ws.iter_rows(values_only=True))
        headers = [h for h in rows[0]]
        for row in rows[1:]:
            data = dict(zip(headers,row))
            dni = (data.get("dni") or "").strip() if data.get("dni") else None
            if dni == "None": dni = None
            if dni:
                cur.execute("SELECT id FROM shooters WHERE dni = ?", (dni,))
                r = cur.fetchone()
            else:
                r = None
            if r:
                cur.execute("""UPDATE shooters SET nombre=?, club=?, categoria=?, licencia=?, pais=? WHERE id=?""",
                            (data.get("nombre"), data.get("club"), data.get("categoria"), data.get("licencia"), data.get("pais"), r['id']))
            else:
                cur.execute("""INSERT OR IGNORE INTO shooters (nombre,club,categoria,licencia,pais,dni) VALUES (?,?,?,?,?,?)""",
                            (data.get("nombre"), data.get("club"), data.get("categoria"), data.get("licencia"), data.get("pais"), dni))
                created += 1
        conn.commit()
    else:
        raise RuntimeError("Necesitas openpyxl para leer .xlsx o proporcionar un CSV alternativo")
    conn.close()
    print(f"Tiradores importados/actualizados: {created}")
    return created

def import_inscriptions_from_excel(path, competition_name):
    conn = get_conn()
    cur = conn.cursor()
    # asegurar competición
    cur.execute("SELECT id FROM competitions WHERE nombre = ?", (competition_name,))
    comp = cur.fetchone()
    if not comp:
        cur.execute("INSERT INTO competitions (nombre) VALUES (?)", (competition_name,))
        comp_id = cur.lastrowid
    else:
        comp_id = comp['id']
    created = 0
    if EXCEL_AVAILABLE and path.lower().endswith('.xlsx'):
        wb = load_workbook(path, read_only=True)
        if "Inscripciones" not in wb.sheetnames:
            print(\"Hoja 'Inscripciones' no encontrada en el fichero.\")
            return 0
        ws = wb["Inscripciones"]
        rows = list(ws.iter_rows(values_only=True))
        headers = [h for h in rows[0]]
        for row in rows[1:]:
            data = dict(zip(headers,row))
            shooter_dni = (data.get("shooter_dni") or "").strip() if data.get("shooter_dni") else None
            if not shooter_dni:
                continue
            cur.execute("SELECT id FROM shooters WHERE dni = ?", (shooter_dni,))
            shooter = cur.fetchone()
            if not shooter:
                print(f\"Aviso: tirador con DNI {shooter_dni} no encontrado. Saltando.\")
                continue
            squad_name = data.get("squad") or "Sin escuadra"
            # buscar o crear escuadra
            cur.execute("SELECT id FROM squads WHERE competition_id=? AND nombre=?",(comp_id,squad_name))
            s = cur.fetchone()
            if not s:
                cur.execute("INSERT INTO squads (competition_id,nombre) VALUES (?,?)",(comp_id,squad_name))
                squad_id = cur.lastrowid
            else:
                squad_id = s['id']
            dorsal = data.get("dorsal")
            cur.execute("INSERT INTO entries (competition_id, shooter_id, squad_id, dorsal) VALUES (?,?,?,?)",
                        (comp_id, shooter['id'], squad_id, str(dorsal) if dorsal is not None else None))
            created += 1
        conn.commit()
    else:
        raise RuntimeError("Necesitas openpyxl para leer .xlsx o proporcionar un CSV alternativo")
    conn.close()
    print(f"Inscripciones importadas: {created}")
    return created

def import_results_from_excel(path, competition_name):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM competitions WHERE nombre = ?", (competition_name,))
    comp = cur.fetchone()
    if not comp:
        raise RuntimeError("Competición no encontrada: " + competition_name)
    comp_id = comp['id']
    created = 0
    if EXCEL_AVAILABLE and path.lower().endswith('.xlsx'):
        wb = load_workbook(path, read_only=True)
        if "Resultados" not in wb.sheetnames:
            print(\"Hoja 'Resultados' no encontrada en el fichero.\")
            return 0
        ws = wb["Resultados"]
        rows = list(ws.iter_rows(values_only=True))
        headers = [h for h in rows[0]]
        for row in rows[1:]:
            data = dict(zip(headers,row))
            dorsal = str(data.get("dorsal")) if data.get("dorsal") is not None else None
            cur.execute("SELECT id FROM entries WHERE competition_id=? AND dorsal=?", (comp_id, dorsal))
            entry = cur.fetchone()
            if not entry:
                print(f\"Aviso: entry con dorsal={dorsal} no encontrada en competición {competition_name}. Saltando.\")
                continue
            round_number = int(data.get("round") or 1)
            hits = int(data.get("hits") or 0)
            misses = int(data.get("misses") or 0)
            score = int(data.get("score") or hits)
            detail = data.get("detalle")
            ts = datetime.datetime.utcnow().isoformat()
            cur.execute("""INSERT INTO round_results (entry_id, round_number, hits, misses, score, detail, timestamp)
                           VALUES (?,?,?,?,?,?,?)""", (entry['id'], round_number, hits, misses, score, str(detail), ts))
            created += 1
        conn.commit()
    else:
        raise RuntimeError("Necesitas openpyxl para leer .xlsx o proporcionar un CSV alternativo")
    conn.close()
    print(f"Resultados importados: {created}")
    return created

def export_results_to_excel(competition_name, out_path):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM competitions WHERE nombre = ?", (competition_name,))
    comp = cur.fetchone()
    if not comp:
        raise RuntimeError(\"Competición no encontrada: \" + competition_name)
    comp_id = comp['id']
    # recolectar datos
    cur.execute(\"\"\"SELECT e.id as entry_id, e.dorsal, s.nombre as shooter_nombre, s.club, s.categoria
                   FROM entries e JOIN shooters s ON e.shooter_id = s.id WHERE e.competition_id = ?\"\"\", (comp_id,))
    entries = cur.fetchall()
    rows = []
    for e in entries:
        cur.execute("SELECT SUM(hits) as total_hits, SUM(score) as total_score FROM round_results WHERE entry_id = ?", (e['entry_id'],))
        totals = cur.fetchone()
        rows.append({
            "dorsal": e['dorsal'],
            "nombre": e['shooter_nombre'],
            "club": e['club'],
            "categoria": e['categoria'],
            "total_hits": totals['total_hits'] if totals['total_hits'] is not None else 0,
            "total_score": totals['total_score'] if totals['total_score'] is not None else 0
        })
    if EXCEL_AVAILABLE and out_path.lower().endswith('.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"
        ws.append(["dorsal","nombre","club","categoria","total_hits","total_score"])
        for r in rows:
            ws.append([r['dorsal'], r['nombre'], r['club'], r['categoria'], r['total_hits'], r['total_score']])
        wb.save(out_path)
        print(f"Resultados exportados a {out_path} ({len(rows)} filas)")
    else:
        # fallback CSV
        import csv
        csv_path = out_path if out_path.lower().endswith('.csv') else out_path + '.csv'
        with open(csv_path,'w',newline='',encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(["dorsal","nombre","club","categoria","total_hits","total_score"])
            for r in rows:
                w.writerow([r['dorsal'], r['nombre'], r['club'], r['categoria'], r['total_hits'], r['total_score']])
        print(f"Resultados exportados a CSV {csv_path} ({len(rows)} filas)")
    conn.close()
    return len(rows)

# ------------------ Ranking simple ------------------
def compute_rankings(competition_name, by_category=True):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM competitions WHERE nombre = ?", (competition_name,))
    comp = cur.fetchone()
    if not comp:
        raise RuntimeError(\"Competición no encontrada: \" + competition_name)
    comp_id = comp['id']
    cur.execute(\"\"\"SELECT e.id as entry_id, e.dorsal, s.nombre as shooter_nombre, s.club, s.categoria
                   FROM entries e JOIN shooters s ON e.shooter_id = s.id WHERE e.competition_id = ?\"\"\", (comp_id,))
    entries = cur.fetchall()
    standings = []
    for e in entries:
        cur.execute("SELECT SUM(hits) as total_hits, SUM(score) as total_score FROM round_results WHERE entry_id = ?", (e['entry_id'],))
        totals = cur.fetchone()
        standings.append({
            "dorsal": e['dorsal'],
            "nombre": e['shooter_nombre'],
            "club": e['club'],
            "categoria": e['categoria'] or "Sin categoría",
            "total_hits": totals['total_hits'] if totals['total_hits'] is not None else 0,
            "total_score": totals['total_score'] if totals['total_score'] is not None else 0
        })
    # ordenar por score desc, hits desc
    standings_sorted = sorted(standings, key=lambda x: (-x['total_score'], -x['total_hits']))
    if by_category:
        grouped = {}
        for s in standings_sorted:
            grouped.setdefault(s['categoria'], []).append(s)
        conn.close()
        return grouped
    conn.close()
    return standings_sorted

# ------------------ Sincronización (stub) ------------------
def push_results_http(competition_name, api_url, api_key=None):
    try:
        import requests
    except Exception:
        raise RuntimeError(\"Para sincronización HTTP instala 'requests' (pip install requests)\")
    data = compute_rankings(competition_name, by_category=False)
    payload = {"competition": competition_name, "standings": data, "timestamp": datetime.datetime.utcnow().isoformat()}
    headers = {'Content-Type': 'application/json'}
    if api_key:
        headers['Authorization'] = f'Bearer {api_key}'
    r = requests.post(api_url, json=payload, headers=headers, timeout=10)
    r.raise_for_status()
    print(\"Sincronización enviada. Respuesta:\", r.status_code, r.text)
    return r.text

# ------------------ GUI (intento) ------------------
def run_gui():
    try:
        from PyQt6.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QPushButton, QFileDialog, QLabel, QMessageBox, QInputDialog
        from PyQt6.QtCore import Qt
    except Exception as e:
        print(\"PyQt6 no está instalado o falló al importarlo. Puedes ejecutar en modo CLI. Para instalar: pip install PyQt6\")
        return

    class MainWindow(QMainWindow):
        def __init__(self):
            super().__init__()
            self.setWindowTitle(\"Fedecat Shooting - Prototipo\")
            self.setMinimumSize(480,240)
            w = QWidget()
            v = QVBoxLayout()
            self.status = QLabel(\"Listo\")
            v.addWidget(self.status)
            btn_templates = QPushButton(\"Crear plantillas (Excel)\")
            btn_templates.clicked.connect(self.create_templates)
            btn_initdb = QPushButton(\"Inicializar BD\")
            btn_initdb.clicked.connect(self.init_db)
            btn_import_shooters = QPushButton(\"Importar Tiradores (.xlsx)\")
            btn_import_shooters.clicked.connect(self.import_shooters)
            btn_import_insc = QPushButton(\"Importar Inscripciones (.xlsx)\")
            btn_import_insc.clicked.connect(self.import_inscriptions)
            btn_import_results = QPushButton(\"Importar Resultados (.xlsx)\")
            btn_import_results.clicked.connect(self.import_results)
            btn_export = QPushButton(\"Exportar Resultados a Excel\")
            btn_export.clicked.connect(self.export_results)
            btn_show_rank = QPushButton(\"Mostrar clasificaciones en consola\")
            btn_show_rank.clicked.connect(self.show_rankings)
            v.addWidget(btn_templates)
            v.addWidget(btn_initdb)
            v.addWidget(btn_import_shooters)
            v.addWidget(btn_import_insc)
            v.addWidget(btn_import_results)
            v.addWidget(btn_export)
            v.addWidget(btn_show_rank)
            w.setLayout(v)
            self.setCentralWidget(w)

        def create_templates(self):
            try:
                create_templates(os.getcwd())
                QMessageBox.information(self, \"Plantillas\", \"Plantillas creadas en el directorio actual.\")
            except Exception as e:
                QMessageBox.warning(self, \"Error\", str(e))

        def init_db(self):
            try:
                init_db()
                QMessageBox.information(self, \"DB\", \"Base de datos inicializada.\")
            except Exception as e:
                QMessageBox.warning(self, \"Error\", str(e))

        def import_shooters(self):
            p, _ = QFileDialog.getOpenFileName(self, \"Seleccionar Tiradores.xlsx\", filter=\"Excel Files (*.xlsx)\")
            if p:
                try:
                    n = import_shooters_from_excel(p)
                    QMessageBox.information(self, \"Importar\", f\"Tiradores importados: {n}\")
                except Exception as e:
                    QMessageBox.warning(self, \"Error\", str(e))

        def import_inscriptions(self):
            p, _ = QFileDialog.getOpenFileName(self, \"Seleccionar Inscripciones.xlsx\", filter=\"Excel Files (*.xlsx)\")
            if p:
                comp, ok = QInputDialog.getText(self, \"Competición\", \"Nombre de la competición:\")
                if ok and comp:
                    try:
                        n = import_inscriptions_from_excel(p, comp)
                        QMessageBox.information(self, \"Importar\", f\"Inscripciones importadas: {n}\")
                    except Exception as e:
                        QMessageBox.warning(self, \"Error\", str(e))

        def import_results(self):
            p, _ = QFileDialog.getOpenFileName(self, \"Seleccionar Resultados.xlsx\", filter=\"Excel Files (*.xlsx)\")
            if p:
                comp, ok = QInputDialog.getText(self, \"Competición\", \"Nombre de la competición:\")
                if ok and comp:
                    try:
                        n = import_results_from_excel(p, comp)
                        QMessageBox.information(self, \"Importar\", f\"Resultados importados: {n}\")
                    except Exception as e:
                        QMessageBox.warning(self, \"Error\", str(e))

        def export_results(self):
            comp, ok = QInputDialog.getText(self, \"Competición\", \"Nombre de la competición:\")
            if not (ok and comp):
                return
            p, _ = QFileDialog.getSaveFileName(self, \"Guardar resultados\", filter=\"Excel Files (*.xlsx);;CSV Files (*.csv)\")
            if p:
                try:
                    n = export_results_to_excel(comp, p)
                    QMessageBox.information(self, \"Exportar\", f\"Exportados {n} filas a {p}\")
                except Exception as e:
                    QMessageBox.warning(self, \"Error\", str(e))

        def show_rankings(self):
            comp, ok = QInputDialog.getText(self, \"Competición\", \"Nombre de la competición:\")
            if not (ok and comp):
                return
            try:
                grouped = compute_rankings(comp, by_category=True)
                txt = json.dumps(grouped, ensure_ascii=False, indent=2)
                print(txt)
                QMessageBox.information(self, \"Clasificaciones\", \"Clasificaciones impresas en consola.\")
            except Exception as e:
                QMessageBox.warning(self, \"Error\", str(e))


    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    app.exec()

# ------------------ CLI y parsing ------------------
def main_cli():
    p = argparse.ArgumentParser(description='Fedecat Shooting - Prototipo')
    p.add_argument('--create-templates', action='store_true', help='Crear archivos de plantilla Excel en el directorio actual')
    p.add_argument('--init-db', action='store_true', help='Inicializar la base de datos local')
    p.add_argument('--import-shooters', metavar='FILE', help='Importar Tiradores.xlsx')
    p.add_argument('--import-inscriptions', nargs=2, metavar=('FILE','COMP'), help='Importar Inscripciones.xlsx y asignar a competición')
    p.add_argument('--import-results', nargs=2, metavar=('FILE','COMP'), help='Importar Resultados.xlsx y asignar a competición')
    p.add_argument('--export-results', nargs=2, metavar=('COMP','OUT'), help='Exportar resultados de competición a archivo')
    p.add_argument('--run-gui', action='store_true', help='Intentar ejecutar GUI (PyQt6 requerido)')
    p.add_argument('--push', nargs=2, metavar=('COMP','URL'), help='Enviar resultados a una API HTTP (requests requerido)')
    args = p.parse_args()

    try:
        if args.create_templates:
            create_templates(os.getcwd())
        if args.init_db:
            init_db()
        if args.import_shooters:
            import_shooters_from_excel(args.import_shooters)
        if args.import_inscriptions:
            fp, comp = args.import_inscriptions
            import_inscriptions_from_excel(fp, comp)
        if args.import_results:
            fp, comp = args.import_results
            import_results_from_excel(fp, comp)
        if args.export_results:
            comp, out = args.export_results
            export_results_to_excel(comp, out)
        if args.push:
            comp, url = args.push
            push_results_http(comp, url)
        if args.run_gui:
            run_gui()
        if not any(vars(args).values()):
            p.print_help()
    except Exception as e:
        print('Error:', e)
        traceback.print_exc()

if __name__ == '__main__':
    main_cli()
