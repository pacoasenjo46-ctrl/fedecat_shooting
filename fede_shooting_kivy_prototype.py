#!/usr/bin/env python3
# FedeShooting KivyMD Prototype - single file
# Requirements: kivy, kivymd, openpyxl
# Run: python fede_shooting_kivy_prototype.py
# This is a prototype UI for tablet/mobile. It focuses on core flows: templates, DB, import, ranking, export.
import os, sqlite3, datetime, traceback
from kivy.core.window import Window

# On desktop, set a reasonable window size for preview
Window.size = (900, 600)

DB_PATH = os.path.join(os.path.dirname(__file__), "fede_shooting_kivy.db")

# ---------- Database helpers ----------
def get_conn():
    conn = sqlite3.connect(DB_PATH, detect_types=sqlite3.PARSE_DECLTYPES|sqlite3.PARSE_COLNAMES)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_conn()
    cur = conn.cursor()
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS shooters (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        numero INTEGER,
        nombre TEXT NOT NULL,
        categoria TEXT,
        comunidad TEXT,
        licencia TEXT UNIQUE
    );
    CREATE TABLE IF NOT EXISTS competitions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL UNIQUE,
        fecha_inicio TEXT
    );
    CREATE TABLE IF NOT EXISTS entries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        competition_id INTEGER NOT NULL,
        shooter_id INTEGER NOT NULL,
        dorsal TEXT,
        FOREIGN KEY(competition_id) REFERENCES competitions(id),
        FOREIGN KEY(shooter_id) REFERENCES shooters(id)
    );
    CREATE TABLE IF NOT EXISTS results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        shooter_id INTEGER NOT NULL,
        serie INTEGER,
        hits INTEGER,
        score INTEGER,
        timestamp TEXT
    );
    """)
    conn.commit()
    conn.close()

# ---------- Excel templates ----------
def create_templates(outdir="."):
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise RuntimeError("openpyxl no está instalado. Instala con: pip install openpyxl")
    # Tiradores_v4
    wb = Workbook(); ws = wb.active; ws.title = "Tiradores"
    ws.append(["Nº","Nombre","Categoría","Comunidad / País","Licencia"])
    ws.append([1,"Juan Pérez","Senior","Andalucía","ESP12345"])
    ws.append([2,"María López","Dama","Galicia","ESP54321"])
    ws.append([3,"Carlos Ruiz","Veterano","Valencia","ESP67890"])
    p1 = os.path.join(outdir,"Tiradores_v4.xlsx"); wb.save(p1)
    # Resultados
    wb2 = Workbook(); ws2 = wb2.active; ws2.title = "Resultados"
    ws2.append(["Licencia","Serie 1","Serie 2","Serie 3","Serie 4","Total"])
    ws2.append(["ESP12345",24,23,24,24,95])
    ws2.append(["ESP54321",22,24,23,24,93])
    ws2.append(["ESP67890",20,21,22,20,83])
    p2 = os.path.join(outdir,"Resultados.xlsx"); wb2.save(p2)
    return p1, p2

# ---------- Import / Export ----------
def import_shooters_from_excel(path):
    try:
        from openpyxl import load_workbook
    except Exception:
        raise RuntimeError("openpyxl required")
    wb = load_workbook(path, read_only=True)
    if "Tiradores" not in wb.sheetnames:
        raise RuntimeError("Hoja 'Tiradores' no encontrada")
    ws = wb["Tiradores"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [h for h in rows[0]]
    idx_num = headers.index("Nº") if "Nº" in headers else None
    idx_nombre = headers.index("Nombre")
    idx_categoria = headers.index("Categoría")
    idx_comunidad = headers.index("Comunidad / País") if "Comunidad / País" in headers else None
    idx_lic = headers.index("Licencia") if "Licencia" in headers else None

    conn = get_conn(); cur = conn.cursor()
    created = 0; updated = 0
    for row in rows[1:]:
        numero = row[idx_num] if idx_num is not None else None
        nombre = row[idx_nombre]
        categoria = row[idx_categoria]
        comunidad = row[idx_comunidad] if idx_comunidad is not None else None
        licencia = str(row[idx_lic]) if idx_lic is not None and row[idx_lic] is not None else None
        if not licencia:
            # use nombre+numero fallback but prefer licencia
            licencia = f"X-{nombre}-{numero}" if numero else f"X-{nombre}"
        # upsert by licencia
        cur.execute("SELECT id FROM shooters WHERE licencia = ?", (licencia,))
        r = cur.fetchone()
        if r:
            cur.execute("UPDATE shooters SET numero=?, nombre=?, categoria=?, comunidad=? WHERE id=?",
                        (numero, nombre, categoria, comunidad, r['id']))
            updated += 1
        else:
            cur.execute("INSERT INTO shooters (numero,nombre,categoria,comunidad,licencia) VALUES (?,?,?,?,?)",
                        (numero,nombre,categoria,comunidad,licencia))
            created += 1
    conn.commit(); conn.close()
    return created, updated

def import_results_from_excel(path):
    try:
        from openpyxl import load_workbook
    except Exception:
        raise RuntimeError("openpyxl required")
    wb = load_workbook(path, read_only=True)
    if "Resultados" not in wb.sheetnames:
        raise RuntimeError("Hoja 'Resultados' no encontrada")
    ws = wb["Resultados"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [h for h in rows[0]]
    idx_lic = headers.index("Licencia")
    # optional series columns
    series_indices = []
    for i,h in enumerate(headers):
        if isinstance(h,str) and h.lower().startswith("serie"):
            series_indices.append(i)
    idx_total = headers.index("Total") if "Total" in headers else None

    conn = get_conn(); cur = conn.cursor()
    created = 0
    for row in rows[1:]:
        licencia = str(row[idx_lic]) if row[idx_lic] is not None else None
        if not licencia: continue
        # find shooter id
        cur.execute("SELECT id FROM shooters WHERE licencia = ?", (licencia,))
        shooter = cur.fetchone()
        if not shooter:
            # skip unknown shooter
            continue
        shooter_id = shooter['id']
        # if total present, use it; else sum series
        total = row[idx_total] if idx_total is not None else None
        if total is None and series_indices:
            s = 0
            for si in series_indices:
                s += int(row[si] or 0)
            total = s
        # save as a single result row with serie=0 and score=total for simplicity
        ts = datetime.datetime.utcnow().isoformat()
        cur.execute("INSERT INTO results (shooter_id, serie, hits, score, timestamp) VALUES (?,?,?,?,?)",
                    (shooter_id, 0, total, total, ts))
        created += 1
    conn.commit(); conn.close()
    return created

def compute_rankings():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""SELECT s.id, s.numero, s.nombre, s.categoria, s.comunidad, s.licencia,
                   COALESCE(SUM(r.score),0) as total_score
                   FROM shooters s LEFT JOIN results r ON s.id = r.shooter_id
                   GROUP BY s.id ORDER BY total_score DESC;""")
    rows = cur.fetchall()
    # general
    general = []
    for pos, r in enumerate(rows, start=1):
        general.append({"puesto": pos, "nombre": r["nombre"], "categoria": r["categoria"], "total": r["total_score"], "comunidad": r["comunidad"]})
    # by category
    cat_map = {}
    for g in general:
        cat = g["categoria"] or "Sin categoría"
        cat_map.setdefault(cat, []).append(g)
    conn.close()
    return general, cat_map

def export_classification_to_excel(competition_name="Competición", out_path="Clasificación.xlsx"):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except Exception:
        raise RuntimeError("openpyxl required")
    general, cat_map = compute_rankings()
    wb = Workbook()
    # style definitions (pastel)
    header_fill = PatternFill(start_color="BCDFFB", end_color="BCDFFB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    alt_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin,right=thin,top=thin,bottom=thin)

    # General sheet
    ws = wb.active; ws.title = "General"
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Clasificación Oficial - {competition_name}"
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].font = Font(size=14, bold=True)
    headers = ["Puesto","Nombre","Categoría","Total","Comunidad / País"]
    ws.append(headers)
    for col_idx, val in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = val
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    # data rows
    for i, row in enumerate(general, start=3):
        ws.append([row["puesto"], row["nombre"], row["categoria"], row["total"], row["comunidad"]])
        if i % 2 == 1:
            for c in range(1,6):
                ws.cell(row=i, column=c).fill = alt_fill

    # category sheets
    for cat, rows in cat_map.items():
        ws_c = wb.create_sheet(title=str(cat)[:31])
        ws_c.merge_cells("A1:E1")
        ws_c["A1"] = f"Clasificación - {cat} - {competition_name}"
        ws_c["A1"].alignment = Alignment(horizontal="center")
        ws_c["A1"].font = Font(size=14, bold=True)
        ws_c.append(headers)
        for col_idx, val in enumerate(headers, start=1):
            cell = ws_c.cell(row=2, column=col_idx)
            cell.value = val
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for i, r in enumerate(rows, start=3):
            ws_c.append([r["puesto"], r["nombre"], r["categoria"], r["total"], r["comunidad"]])
            if i % 2 == 1:
                for c in range(1,6):
                    ws_c.cell(row=i, column=c).fill = alt_fill

    wb.save(out_path)
    return out_path

# ---------- KivyMD UI ----------
from kivymd.app import MDApp
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.toolbar import MDTopAppBar
from kivymd.uix.button import MDRaisedButton, MDFlatButton
from kivymd.uix.label import MDLabel
from kivymd.uix.dialog import MDDialog
from kivymd.uix.filemanager import MDFileManager
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.metrics import dp
from kivy.utils import platform

class MainScreen(MDBoxLayout):
    def __init__(self, **kv):
        super().__init__(orientation="vertical", spacing=10, padding=10)
        self.app = MDApp.get_running_app()
        self.toolbar = MDTopAppBar(title="FedeShooting - Prototipo", elevation=6)
        self.add_widget(self.toolbar)
        self.body = MDBoxLayout(orientation="vertical", spacing=8, size_hint_y=None)
        self.body.bind(minimum_height=self.body.setter('height'))
        # buttons
        btn_row = MDBoxLayout(size_hint_y=None, height=dp(48), spacing=8)
        btn_row.add_widget(MDRaisedButton(text="Inicializar BD", on_release=self.on_init_db))
        btn_row.add_widget(MDRaisedButton(text="Crear plantillas", on_release=self.on_create_templates))
        btn_row.add_widget(MDRaisedButton(text="Importar Tiradores", on_release=self.on_import_tiradores))
        btn_row.add_widget(MDRaisedButton(text="Importar Resultados", on_release=self.on_import_resultados))
        self.add_widget(btn_row)
        # actions row 2
        btn_row2 = MDBoxLayout(size_hint_y=None, height=dp(48), spacing=8)
        btn_row2.add_widget(MDRaisedButton(text="Calcular Clasificación", on_release=self.on_compute))
        btn_row2.add_widget(MDRaisedButton(text="Exportar Clasificación", on_release=self.on_export))
        btn_row2.add_widget(MDFlatButton(text="Mostrar en consola", on_release=self.on_show_console))
        self.add_widget(btn_row2)
        # info label
        self.info = MDLabel(text="Estado: listo", size_hint_y=None, height=dp(30))
        self.add_widget(self.info)
        # file manager
        self.file_manager = MDFileManager(exit_manager=self.exit_manager, select_path=self.select_path)
        self._fm_callback = None

    def on_init_db(self, *a):
        try:
            init_db()
            self.info.text = "DB inicializada"
        except Exception as e:
            self.info.text = "Error al inicializar DB: " + str(e)

    def on_create_templates(self, *a):
        try:
            p1,p2 = create_templates(os.getcwd())
            self.info.text = f"Plantillas creadas: {p1}, {p2}"
        except Exception as e:
            self.info.text = "Error plantillas: " + str(e)

    def open_file_manager(self, callback):
        self._fm_callback = callback
        # start at user home for better UX on mobile
        start_dir = os.path.expanduser("~")
        self.file_manager.show(start_dir)

    def select_path(self, path):
        # when a file is chosen, call the stored callback
        if self._fm_callback:
            cb = self._fm_callback
            self._fm_callback = None
            cb(path)
        self.exit_manager()

    def exit_manager(self, *args):
        self.file_manager.close()

    def on_import_tiradores(self, *a):
        self.open_file_manager(self._import_tiradores_cb)

    def _import_tiradores_cb(self, path):
        try:
            created, updated = import_shooters_from_excel(path)
            self.info.text = f"Tiradores: {created} creados, {updated} actualizados"
        except Exception as e:
            self.info.text = "Error import tiradores: " + str(e)

    def on_import_resultados(self, *a):
        self.open_file_manager(self._import_resultados_cb)

    def _import_resultados_cb(self, path):
        try:
            created = import_results_from_excel(path)
            self.info.text = f"Resultados importados: {created}"
        except Exception as e:
            self.info.text = "Error import resultados: " + str(e)

    def on_compute(self, *a):
        try:
            gen, cat = compute_rankings()
            self.info.text = f"Clasificación: {len(gen)} tiradores procesados"
        except Exception as e:
            self.info.text = "Error compute: " + str(e)

    def on_export(self, *a):
        try:
            out = os.path.join(os.getcwd(), "Clasificacion.xlsx")
            export_classification_to_excel("Competición de Prueba", out)
            self.info.text = "Clasificación exportada: " + out
        except Exception as e:
            self.info.text = "Error export: " + str(e)

    def on_show_console(self, *a):
        try:
            gen, cat = compute_rankings()
            import json, sys
            print(json.dumps({"general":gen,"by_category":cat}, ensure_ascii=False, indent=2))
            self.info.text = "Clasificación impresa en consola"
        except Exception as e:
            self.info.text = "Error mostrar: " + str(e)

class FedeApp(MDApp):
    def build(self):
        init_db()
        return MainScreen()

if __name__ == "__main__":
    FedeApp().run()
